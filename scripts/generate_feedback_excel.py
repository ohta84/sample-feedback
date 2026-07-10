#!/usr/bin/env python3
"""
Bカート「サンプル評価回収CSV」+「受注CSV」-> メーカー別サンプルフィードバックExcel 生成スクリプト (v2)

使い方:
  python generate_feedback_excel_v2.py \
      --eval-csv サンプル評価回収.csv \
      --order-csv bcart_order.csv \
      --maker-name "○○" \
      --output "○○様サンプルフィードバック.xlsx"

--maker-code は省略可。複数コードはカンマ区切り: --maker-code "121,122"
"""
import argparse
import os
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MAKER_CODES_CSV = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "maker_codes.csv")


def load_maker_code_map(path=MAKER_CODES_CSV):
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    return dict(zip(df.iloc[:, 0].str.strip(), df.iloc[:, 1].str.strip()))


EXCLUDE_COMPANY_KEYWORDS = ["ナゴミヤ", "なごみや"]

HEADER_KEYWORDS = {
    "納品日": ["納品日"],
    "業態": ["業態"],
    "会社名": ["会社名"],
    "都道府県": ["配送先 都道府県", "配送先都道府県", "都道府県"],
    "商品名": ["商品名"],
    "セット名": ["セット名"],
    "受注番号": ["受注番号"],
    "サンプル用途": ["サンプル用途"],
    "評価結果": ["サンプル評価結果"],
    "採用数": ["評価都合／採用数", "評価都合/採用数"],
    "受注数": ["受注数"],
    "小計": ["小計"],
}

FONT_NAME = "Arial"
GREY_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
GREEN_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
THIN_GREY = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

COLUMN_WIDTHS = {
    "A": 14.6, "B": 15.1, "C": 31.6, "D": 11.9,
    "E": 37.0, "F": 35.2, "G": 71.4, "H": 14.0,
}


def read_csv_flexible(path):
    for enc in ("cp932", "utf-8-sig", "utf-8", "shift_jis"):
        try:
            return pd.read_csv(path, encoding=enc, dtype=str, keep_default_na=False)
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise RuntimeError(f"{path} の文字コードを判定できませんでした")


def find_column(df, keywords):
    for kw in keywords:
        for col in df.columns:
            if kw in str(col):
                return col
    return None


def extract_maker_code(product_name):
    m = re.match(r"^\s*\[(\d+)\]", str(product_name))
    return m.group(1) if m else None


def parse_date(value):
    value = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def count_keyword(text, kw):
    return str(text).count(kw)


def to_number(value):
    s = str(value).strip().replace(",", "")
    if s in ("", "nan", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


CORP_PREFIXES = ["株式会社", "有限会社", "合同会社", "（株）", "(株)", "（有）", "(有)"]


def clean_company_name(name):
    s = str(name).strip()
    for p in CORP_PREFIXES:
        s = s.replace(p, "")
    return s.strip()


def detect_maker_code(series_of_product_names):
    codes = [extract_maker_code(p) for p in series_of_product_names]
    codes = [c for c in codes if c]
    if not codes:
        return None
    return pd.Series(codes).value_counts().idxmax()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--eval-csv", required=True, help="サンプル評価回収CSV")
    ap.add_argument("--order-csv", required=True, help="Bカート受注CSV")
    ap.add_argument("--maker-name", default=None, help="省略時はCSVから自動検出したメーカーコードとdata/maker_codes.csvから自動解決")
    ap.add_argument("--maker-code", default=None, help="省略時はCSVの商品名から自動検出。カンマ区切りで複数指定可")
    ap.add_argument("--output", default=None, help="省略時は '{コード}{会社名}様サンプルフィードバック.xlsx' を自動生成")
    args = ap.parse_args()

    df_eval = read_csv_flexible(args.eval_csv)
    df_order = read_csv_flexible(args.order_csv)

    o_date = find_column(df_order, HEADER_KEYWORDS["納品日"])
    o_type = find_column(df_order, HEADER_KEYWORDS["業態"])
    o_company = find_column(df_order, HEADER_KEYWORDS["会社名"])
    o_pref = find_column(df_order, HEADER_KEYWORDS["都道府県"])
    o_product = find_column(df_order, HEADER_KEYWORDS["商品名"])
    o_set = find_column(df_order, HEADER_KEYWORDS["セット名"])
    o_orderno = find_column(df_order, HEADER_KEYWORDS["受注番号"])
    o_qty = find_column(df_order, HEADER_KEYWORDS["受注数"])
    o_subtotal = find_column(df_order, HEADER_KEYWORDS["小計"])

    e_date = find_column(df_eval, HEADER_KEYWORDS["納品日"])
    e_company = find_column(df_eval, HEADER_KEYWORDS["会社名"])
    e_product = find_column(df_eval, HEADER_KEYWORDS["商品名"])
    e_orderno = find_column(df_eval, HEADER_KEYWORDS["受注番号"])
    e_usage = find_column(df_eval, HEADER_KEYWORDS["サンプル用途"])
    e_result = find_column(df_eval, HEADER_KEYWORDS["評価結果"])
    e_adopt = find_column(df_eval, HEADER_KEYWORDS["採用数"])

    missing = [n for n, c in [
        ("受注CSV:納品日", o_date), ("受注CSV:会社名", o_company), ("受注CSV:商品名", o_product),
        ("受注CSV:セット名", o_set), ("受注CSV:受注番号", o_orderno),
        ("評価CSV:会社名", e_company), ("評価CSV:商品名", e_product), ("評価CSV:受注番号", e_orderno),
    ] if c is None]
    if missing:
        sys.stderr.write("以下の列が見つかりませんでした: " + ", ".join(missing) + "\n")
        sys.exit(1)

    # メーカーコードの決定（未指定ならCSVの商品名から自動検出）
    if args.maker_code:
        maker_code_for_name = args.maker_code.split(",")[0].strip()
    else:
        detected = detect_maker_code(df_order[o_product])
        if detected:
            args.maker_code = detected
            maker_code_for_name = detected
        else:
            maker_code_for_name = None

    # メーカー名の決定（未指定ならコード表から自動解決、法人格は除去）
    if not args.maker_name:
        if not maker_code_for_name:
            sys.stderr.write("メーカーコードをCSVから自動検出できませんでした。--maker-name か --maker-code を指定してください。\n")
            sys.exit(1)
        code_map = load_maker_code_map()
        resolved = code_map.get(maker_code_for_name)
        if not resolved:
            sys.stderr.write(f"メーカーコード {maker_code_for_name} が data/maker_codes.csv に見つかりません。--maker-name を指定してください。\n")
            sys.exit(1)
        args.maker_name = f"{maker_code_for_name}{clean_company_name(resolved)}"

    if not args.output:
        args.output = f"{args.maker_name}様サンプルフィードバック.xlsx"

    df_order["_is_sample"] = df_order[o_set].astype(str).str.contains("サンプル", na=False)

    # 受注CSVから 業態・都道府県 を 受注番号+商品名 で引けるようにする
    lookup = df_order[[o_orderno, o_product, o_type, o_pref, o_set]].drop_duplicates(
        subset=[o_orderno, o_product]
    )
    lookup = lookup.rename(columns={o_orderno: "_key_orderno", o_product: "_key_product",
                                     o_type: "_業態", o_pref: "_都道府県", o_set: "_セット名"})

    df = df_eval.rename(columns={e_orderno: "_key_orderno", e_product: "_key_product"})
    df = df.merge(lookup, on=["_key_orderno", "_key_product"], how="left")

    # 評価CSVには通常注文行も混在しているため、セット名でサンプル行のみに絞り込む
    df = df[df["_セット名"].astype(str).str.contains("サンプル", na=False)]

    # 自社テスト除外
    df = df[~df[e_company].astype(str).apply(
        lambda v: any(kw in v for kw in EXCLUDE_COMPANY_KEYWORDS)
    )]
    df_order = df_order[~df_order[o_company].astype(str).apply(
        lambda v: any(kw in v for kw in EXCLUDE_COMPANY_KEYWORDS)
    )]

    # メーカーコードで絞り込み
    if args.maker_code:
        codes = [c.strip() for c in args.maker_code.split(",") if c.strip()]
        df = df[df["_key_product"].apply(lambda p: extract_maker_code(p) in codes)]
        df_order = df_order[df_order[o_product].apply(lambda p: extract_maker_code(p) in codes)]

    if df.empty:
        sys.stderr.write("絞り込み後、評価データが0件でした。メーカーコードを確認してください。\n")
        sys.exit(1)

    # 通常注文をした会社（受注CSV全体から。セット名にサンプルを含まない行がある会社）
    companies_with_order = set(
        df_order.loc[~df_order["_is_sample"], o_company].astype(str)
    )

    df["_company"] = df[e_company].astype(str)
    df["_date_parsed"] = df[e_date].apply(parse_date) if e_date else None

    provided_companies = sorted(set(df["_company"]))
    n_provided = len(provided_companies)
    n_converted = sum(1 for c in provided_companies if c in companies_with_order)
    rate = (n_converted / n_provided) if n_provided else 0

    all_dates = [d for d in df["_date_parsed"] if d is not None] if e_date else []
    period_str = f"{min(all_dates):%Y/%m/%d}〜{max(all_dates):%Y/%m/%d}" if all_dates else ""

    df_sorted = df.sort_values(["_company", "_date_parsed"])

    groups = []
    for (company, date_val), g in df_sorted.groupby(["_company", e_date], sort=False):
        groups.append((company, date_val, g))
    groups.sort(key=lambda x: (x[0], parse_date(x[1]) or datetime.max))

    # ---------------- 業態別分析用の集計 ----------------
    company_type = {}
    for _, row in df.iterrows():
        c = row["_company"]
        t = row.get("_業態")
        if c not in company_type and pd.notna(t) and str(t).strip():
            company_type[c] = str(t).strip()

    biz_stats = {}
    for c in provided_companies:
        t = company_type.get(c, "(不明)")
        s = biz_stats.setdefault(t, {"提供": 0, "転換": 0, "OK件数": 0, "NG件数": 0, "検討中件数": 0})
        s["提供"] += 1
        if c in companies_with_order:
            s["転換"] += 1

    for _, row in df.iterrows():
        t = company_type.get(row["_company"], "(不明)")
        text = str(row.get(e_result, "")) if e_result else ""
        biz_stats[t]["OK件数"] += count_keyword(text, "OK")
        biz_stats[t]["NG件数"] += count_keyword(text, "NG")
        biz_stats[t]["検討中件数"] += count_keyword(text, "検討中")

    # ---------------- 受注データ分析（受注CSV全体・通常注文ベース） ----------------
    df_normal = df_order[~df_order["_is_sample"]].copy()
    if o_qty:
        df_normal["_qty"] = df_normal[o_qty].apply(to_number)
    else:
        df_normal["_qty"] = 0.0
    if o_subtotal:
        df_normal["_amount"] = df_normal[o_subtotal].apply(to_number)
    else:
        df_normal["_amount"] = 0.0

    # 業態別 受注動向
    order_biz_stats = {}
    for _, row in df_normal.iterrows():
        t = str(row[o_type]).strip() if o_type and str(row[o_type]).strip() else "(不明)"
        s = order_biz_stats.setdefault(t, {"companies": set(), "件数": 0, "金額": 0.0})
        s["companies"].add(str(row[o_company]))
        s["件数"] += 1
        s["金額"] += row["_amount"]

    # 商品別ランキング
    product_stats = {}
    for _, row in df_normal.iterrows():
        p = str(row[o_product]).strip()
        s = product_stats.setdefault(p, {"companies": set(), "数量": 0.0, "金額": 0.0})
        s["companies"].add(str(row[o_company]))
        s["数量"] += row["_qty"]
        s["金額"] += row["_amount"]

    # ---------------- Excel生成 ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "サンプルフィードバック"

    def set_row(row_idx, text, bold=False, size=11, color=None):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
        cell.alignment = Alignment(wrap_text=False)
        return cell

    set_row(1, f"{args.maker_name}様サンプルフィードバック", bold=True, size=18)
    set_row(2, f"サンプル提供社数：{n_provided}社")
    set_row(3, f"注文転換社数：{n_converted}社")
    cell4 = set_row(4, f"注文転換率：{rate*100:.1f}%", bold=True, color="FF0000")

    set_row(6, period_str)
    updater_cell = ws.cell(row=6, column=8, value="更新日時：なごみや太田")
    updater_cell.font = Font(name=FONT_NAME, size=11)
    updater_cell.alignment = Alignment(wrap_text=False, horizontal="right")

    headers = ["納品日", "業態", "会社名", "配送先都道府県", "商品名", "サンプル用途", "サンプル評価結果", "注文有無"]
    header_row = 7
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = GREY_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER_ALL

    row_ptr = header_row + 1
    for company, date_val, g in groups:
        products = list(g["_key_product"])
        n_rows = max(len(products), 1)
        start_row = row_ptr
        end_row = row_ptr + n_rows - 1

        d_parsed = parse_date(date_val)
        date_display = f"{d_parsed:%Y/%m/%d}" if d_parsed else date_val

        has_order = company in companies_with_order
        order_text = "✅ 注文あり" if has_order else "－ なし"

        first = g.iloc[0]
        result_text = str(first[e_result]).strip() if e_result else ""
        adopt_text = str(first[e_adopt]).strip() if e_adopt else ""
        if adopt_text and result_text:
            g_value = f"評価都合／採用数：{adopt_text}\n{result_text}"
        else:
            g_value = result_text or adopt_text

        merged_values = {
            "A": date_display,
            "B": first.get("_業態", ""),
            "C": company,
            "D": first.get("_都道府県", ""),
            "F": str(first[e_usage]).strip() if e_usage else "",
            "G": g_value,
            "H": order_text,
        }

        for col_letter, value in merged_values.items():
            col_idx = "ABCDEFGH".index(col_letter) + 1
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            cell.font = Font(name=FONT_NAME, size=11,
                              bold=(col_letter == "H" and has_order),
                              color=("375623" if (col_letter == "H" and has_order)
                                     else ("999999" if col_letter == "H" else None)))
            if col_letter == "H" and has_order:
                cell.fill = GREEN_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                                end_row=end_row, end_column=col_idx)
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=col_idx).border = BORDER_ALL

        for offset, product in enumerate(products):
            r = start_row + offset
            cell = ws.cell(row=r, column=5, value=product)
            cell.font = Font(name=FONT_NAME, size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER_ALL

        row_ptr = end_row + 1

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ---------------- シート2: 業態別分析 ----------------
    ws2 = wb.create_sheet("業態別分析")
    headers2 = ["業態", "提供社数", "転換社数", "転換率", "OK件数(概算)", "NG件数(概算)", "検討中件数(概算)"]
    for i, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = GREY_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER_ALL

    sorted_biz = sorted(biz_stats.items(), key=lambda kv: kv[1]["提供"], reverse=True)
    for r, (biz, s) in enumerate(sorted_biz, start=2):
        biz_rate = (s["転換"] / s["提供"]) if s["提供"] else 0
        values = [biz, s["提供"], s["転換"], biz_rate, s["OK件数"], s["NG件数"], s["検討中件数"]]
        for i, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=i, value=v)
            cell.font = Font(name=FONT_NAME, size=11)
            cell.border = BORDER_ALL
            if i == 4:
                cell.number_format = "0.0%"

    for col_letter, width in {"A": 16, "B": 10, "C": 10, "D": 10, "E": 12, "F": 12, "G": 14}.items():
        ws2.column_dimensions[col_letter].width = width

    note = ws2.cell(row=len(sorted_biz) + 3, column=1,
                     value="※OK/NG/検討中件数は評価コメント内の文字列を数えた概算値です（1コメントに複数商品分の評価が含まれる場合があります）")
    note.font = Font(name=FONT_NAME, size=9, italic=True)

    # ---------------- シート3: 受注データ分析 ----------------
    ws3 = wb.create_sheet("受注データ分析")

    ws3.cell(row=1, column=1, value="業態別 受注動向（通常注文ベース）").font = Font(name=FONT_NAME, size=13, bold=True)
    headers3a = ["業態", "受注社数", "受注件数", "受注金額合計", "平均受注額／社"]
    for i, h in enumerate(headers3a, start=1):
        c = ws3.cell(row=2, column=i, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = GREY_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER_ALL

    sorted_order_biz = sorted(order_biz_stats.items(), key=lambda kv: kv[1]["金額"], reverse=True)
    r = 3
    for biz, s in sorted_order_biz:
        n_companies = len(s["companies"])
        avg = (s["金額"] / n_companies) if n_companies else 0
        values = [biz, n_companies, s["件数"], s["金額"], avg]
        for i, v in enumerate(values, start=1):
            cell = ws3.cell(row=r, column=i, value=v)
            cell.font = Font(name=FONT_NAME, size=11)
            cell.border = BORDER_ALL
            if i in (4, 5):
                cell.number_format = "#,##0"
        r += 1

    table2_start = r + 2
    ws3.cell(row=table2_start, column=1, value="商品別ランキング（通常注文・金額順、上位10件）").font = Font(name=FONT_NAME, size=13, bold=True)
    headers3b = ["商品名", "受注数量合計", "受注金額合計", "購入社数"]
    for i, h in enumerate(headers3b, start=1):
        c = ws3.cell(row=table2_start + 1, column=i, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = GREY_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER_ALL

    sorted_products = sorted(product_stats.items(), key=lambda kv: kv[1]["金額"], reverse=True)[:10]
    r = table2_start + 2
    for product, s in sorted_products:
        values = [product, s["数量"], s["金額"], len(s["companies"])]
        for i, v in enumerate(values, start=1):
            cell = ws3.cell(row=r, column=i, value=v)
            cell.font = Font(name=FONT_NAME, size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = BORDER_ALL
            if i == 3:
                cell.number_format = "#,##0"
        r += 1

    for col_letter, width in {"A": 45, "B": 14, "C": 16, "D": 12, "E": 16}.items():
        ws3.column_dimensions[col_letter].width = width

    note3 = ws3.cell(row=r + 2, column=1,
                      value="※このシートはサンプル注文を除いた「通常注文」のみを集計しています")
    note3.font = Font(name=FONT_NAME, size=9, italic=True)

    wb.save(args.output)
    print(f"OK: {args.output} (提供社数={n_provided}, 転換社数={n_converted}, 転換率={rate*100:.1f}%)")


if __name__ == "__main__":
    main()
